"""Local OpenPyXL / CSV export for department cohorts."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.sheet_template import ROW3_VALUES, ROW4_VALUES, format_date_banner

NAVY_FILL = PatternFill("solid", fgColor="1F497D")
DARK_FILL = PatternFill("solid", fgColor="2C333F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DATE_FONT = Font(color="FFFFFF", bold=True, italic=True)
STRIPE_FILL = PatternFill("solid", fgColor="F2F2F2")
STATUS_FILLS = {
    "Top-K Pass": PatternFill("solid", fgColor="C6EFCE"),
    "Review": PatternFill("solid", fgColor="FFEB9C"),
    "Fail": PatternFill("solid", fgColor="FFC7CE"),
}


def cohort_to_dataframe(rows: list[dict[str, Any]]) -> pd.DataFrame:
    records = []
    for row in rows:
        records.append(
            {
                "INTERVIEW TIME": row.get("interview_time", ""),
                "NAME": row.get("name", ""),
                "1st Choice": row.get("first_choice", ""),
                "2nd Choice": row.get("second_choice", ""),
                "Meeting Link": row.get("meeting_link", "N/A"),
                "Notes": row.get("notes", ""),
                "STATUS": row.get("status", ""),
            }
        )
    return pd.DataFrame(records)


def export_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    df = cohort_to_dataframe(rows)
    return df.to_csv(index=False).encode("utf-8")


def _style_header_block(ws) -> None:
    """Match department sheet layout: 4 header rows, data from row 5."""
    date_label = format_date_banner(datetime.now(timezone.utc))

    for col in range(1, 8):
        ws.cell(row=1, column=col).fill = DARK_FILL

    ws["A2"] = date_label

    for col_idx, val in enumerate(ROW3_VALUES, start=1):
        ws.cell(row=3, column=col_idx, value=val)
    for col_idx, val in enumerate(ROW4_VALUES, start=1):
        ws.cell(row=4, column=col_idx, value=val)

    ws.merge_cells("A2:G2")
    ws.merge_cells("A3:A4")
    ws.merge_cells("B3:B4")
    ws.merge_cells("C3:D3")
    ws.merge_cells("E3:E4")
    ws.merge_cells("F3:F4")
    ws.merge_cells("G3:G4")

    date_cell = ws["A2"]
    date_cell.fill = DARK_FILL
    date_cell.font = DATE_FONT
    date_cell.alignment = Alignment(horizontal="left", vertical="center")

    row3_cols = (1, 2, 3, 5, 6, 7)  # skip D3 (merged into C3:D3)
    for col_idx in row3_cols:
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in (3, 4):  # 1st / 2nd choice sub-headers
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def export_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cohort"

    _style_header_block(ws)

    for idx, row in enumerate(rows, start=5):
        values = [
            row.get("interview_time", ""),
            row.get("name", ""),
            row.get("first_choice", ""),
            row.get("second_choice", ""),
            row.get("meeting_link", "N/A"),
            row.get("notes", ""),
            row.get("status", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            if idx % 2 == 1:
                cell.fill = STRIPE_FILL
            if col_idx == 7:
                status_fill = STATUS_FILLS.get(str(value))
                if status_fill:
                    cell.fill = status_fill

    for col_idx in range(1, 8):
        letter = get_column_letter(col_idx)
        max_len = 12
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
